# -*- coding: utf-8 -*-
"""
搜索结果（与 ``pc_search_export`` / 库内 ``JdJobSearchRow`` 一致的中文宽表）经大模型补充**品牌**、**规格摘要**、**分类（叶子类目短语）**，导出多工作表 xlsx。

多批请求默认 **并行**（``ThreadPoolExecutor``），大数据量时可明显缩短墙钟时间；若网关限流可减小
环境变量 ``PC_SEARCH_LLM_MAX_WORKERS``（默认 8）或查询参数 ``max_workers=2``。

CLI 见 ``pipeline.demos.pc_search_llm_brand_spec_excel``；线上经 ``JobDatasetSearchLlmXlsxView`` 下载。
"""
from __future__ import annotations

import json
import logging
import os
from collections import Counter
from concurrent.futures import ThreadPoolExecutor, as_completed
from io import BytesIO
from pathlib import Path
from typing import Any

from openpyxl import Workbook

from pipeline.export_job import _search_row_csv_dict
from pipeline.llm.llm_client import call_llm
from pipeline.models import JdJobSearchRow, PipelineJob
from pipeline.dataset_nonempty import nonempty_search_keys_for_job, search_export_headers

logger = logging.getLogger(__name__)

LLM_EXTRA_HEADERS = ("LLM品牌", "LLM规格摘要", "LLM分类", "LLM置信度", "LLM备注")

SYSTEM_PROMPT = """你是电商商品信息抽取助手，根据搜索导出行的「标题」「规格属性」「店铺名」，以及可选的「类目线索 category_hint」，推断：
- brand：品牌名（用常见标准写法，如标题中仅为系列名则结合店铺名推断；仍无法判断则空字符串）
- spec_summary：一句规格摘要（净含量/ml、g、套装件数、容量等；没有则空字符串）
- category：叶子类目或细分品类短标签（便于统计归类，例如「面部精华」「保湿凝胶」；与平台原始类目不必逐字一致；无从判断则空字符串；若 category_hint 有合理信息可采纳，与标题明显冲突时以标题为准）
- confidence：high / medium / low
- notes：可选，极短说明（可为空字符串）

必须只输出一个合法 JSON 数组，不要 Markdown、不要代码围栏、不要中英文解释。
数组元素字段：row_index（整数，与输入一致）, brand, spec_summary, category, confidence, notes。"""


def llm_payload_item_from_row(r: dict[str, Any], row_index: int) -> dict[str, Any]:
    """兼容英文内部键（入库 API）与中文 CSV 表头。"""
    sku = (
        r.get("sku_id")
        or r.get("SKU")
        or r.get("主商品ID")
        or r.get("item_id")
        or ""
    )
    title = r.get("title") or r.get("标题") or ""
    spec = r.get("attributes") or r.get("规格属性") or ""
    shop = r.get("shop_name") or r.get("店铺名") or ""
    kw = r.get("keyword") or r.get("搜索词") or ""
    cat_hint = (
        r.get("leaf_category")
        or r.get("类目")
        or r.get("leafCategory")
        or ""
    )
    return {
        "row_index": row_index,
        "sku": str(sku).strip(),
        "title": str(title).strip(),
        "spec": str(spec).strip(),
        "shop": str(shop).strip(),
        "search_keyword": str(kw).strip(),
        "category_hint": str(cat_hint).strip(),
    }


def llm_json_payload_slice(start: int, slice_rows: list[dict[str, Any]]) -> str:
    items = [llm_payload_item_from_row(r, start + i) for i, r in enumerate(slice_rows)]
    # 紧凑 JSON，省 token、间接允许略大的 batch_size
    return json.dumps(items, ensure_ascii=False, separators=(",", ":"))


def parse_llm_items(raw: str) -> list[dict[str, Any]]:
    t = (raw or "").strip()
    try:
        data = json.loads(t)
    except json.JSONDecodeError:
        s, e = t.find("["), t.rfind("]")
        if s < 0 or e <= s:
            raise ValueError(
                f"模型输出无法解析为 JSON 数组，前 400 字：{t[:400]!r}"
            ) from None
        data = json.loads(t[s : e + 1])
    if not isinstance(data, list):
        raise ValueError("模型输出应为 JSON 数组")
    return [x for x in data if isinstance(x, dict)]


def classify_batch(user_json_block: str) -> list[dict[str, Any]]:
    user = f"""请处理下列商品行，为每一行填写 brand、spec_summary、category、confidence、notes，并原样返回 row_index。

输入：
{user_json_block}
"""
    raw = call_llm(SYSTEM_PROMPT, user, temperature=0.1)
    return parse_llm_items(raw)


def _default_max_workers() -> int:
    raw = (os.environ.get("PC_SEARCH_LLM_MAX_WORKERS") or "").strip()
    if raw:
        try:
            return max(1, min(int(raw), 32))
        except ValueError:
            pass
    return 8


def _run_single_batch(start: int, chunk: list[dict[str, Any]]) -> tuple[int, int, list[dict[str, Any]]]:
    """返回 (start, chunk_len, items)。"""
    payload = llm_json_payload_slice(start, chunk)
    items = classify_batch(payload)
    return start, len(chunk), items


def run_llm_classify_by_row_index(
    rows: list[dict[str, Any]],
    *,
    batch_size: int = 40,
    max_workers: int | None = None,
) -> dict[int, dict[str, Any]]:
    n = len(rows)
    batch = max(1, min(int(batch_size), 120))
    workers = _default_max_workers() if max_workers is None else max(1, min(int(max_workers), 32))
    n_chunks = (n + batch - 1) // batch
    by_index: dict[int, dict[str, Any]] = {}

    def _merge_items(start: int, chunk_len: int, items: list[dict[str, Any]]) -> None:
        for it in items:
            try:
                ri = int(it["row_index"])
            except (KeyError, TypeError, ValueError):
                continue
            by_index[ri] = it
        miss = [start + i for i in range(chunk_len) if (start + i) not in by_index]
        if miss:
            logger.warning("pc_search_llm: 本批 start=%s 未返回 row_index: %s", start, miss[:15])

    if workers <= 1 or n_chunks <= 1:
        for idx, start in enumerate(range(0, n, batch)):
            chunk = rows[start : start + batch]
            logger.info(
                "pc_search_llm: 批次 %s/%s 行 %s-%s",
                idx + 1,
                n_chunks,
                start,
                start + len(chunk) - 1,
            )
            _, chunk_len, items = _run_single_batch(start, chunk)
            _merge_items(start, chunk_len, items)
        return by_index

    logger.info(
        "pc_search_llm: 并行 workers=%s batch_size=%s 共 %s 批、%s 行",
        workers,
        batch,
        n_chunks,
        n,
    )
    with ThreadPoolExecutor(max_workers=workers) as ex:
        future_to_start = {}
        for start in range(0, n, batch):
            chunk = rows[start : start + batch]
            fut = ex.submit(_run_single_batch, start, chunk)
            future_to_start[fut] = (start, len(chunk))
        for fut in as_completed(future_to_start):
            start, chunk_len = future_to_start[fut]
            try:
                s_ret, clen_ret, items = fut.result()
            except Exception:
                logger.exception("pc_search_llm: 批次失败 start=%s", start)
                raise
            _merge_items(s_ret, clen_ret, items)
    return by_index


def merge_llm_into_rows(
    headers: list[str],
    rows: list[dict[str, Any]],
    by_index: dict[int, dict[str, Any]],
) -> tuple[list[str], list[dict[str, Any]]]:
    new_headers = list(headers) + [h for h in LLM_EXTRA_HEADERS if h not in headers]
    out: list[dict[str, Any]] = []
    for i, r in enumerate(rows):
        nr = dict(r)
        rec = by_index.get(i, {})
        nr["LLM品牌"] = str(rec.get("brand") or "").strip()
        nr["LLM规格摘要"] = str(rec.get("spec_summary") or "").strip()
        nr["LLM分类"] = str(
            rec.get("category") or rec.get("category_llm") or ""
        ).strip()
        nr["LLM置信度"] = str(rec.get("confidence") or "").strip()
        nr["LLM备注"] = str(rec.get("notes") or "").strip()
        out.append(nr)
    return new_headers, out


def _excel_cell(s: str, max_len: int = 32700) -> str:
    if len(s) <= max_len:
        return s
    return s[: max_len - 12] + "…(已截断)"


def write_llm_workbook_bytes(headers: list[str], merged_rows: list[dict[str, Any]]) -> bytes:
    wb = Workbook()
    ws0 = wb.active
    ws0.title = "源数据"
    ws0.append(headers)
    for r in merged_rows:
        ws0.append([_excel_cell(str(r.get(h, "") or "")) for h in headers])

    unk = "（空）"
    brands = Counter()
    specs = Counter()
    cats = Counter()
    for r in merged_rows:
        b = (r.get("LLM品牌") or "").strip() or unk
        brands[b] += 1
        sp = (r.get("LLM规格摘要") or "").strip() or unk
        specs[sp] += 1
        c = (r.get("LLM分类") or "").strip() or unk
        cats[c] += 1

    ws1 = wb.create_sheet("品牌统计")
    ws1.append(["品牌", "数量"])
    for name, cnt in brands.most_common():
        ws1.append([name, cnt])

    ws2 = wb.create_sheet("规格统计")
    ws2.append(["规格摘要", "数量"])
    for name, cnt in specs.most_common():
        ws2.append([name, cnt])

    ws3 = wb.create_sheet("分类统计")
    ws3.append(["LLM分类", "数量"])
    for name, cnt in cats.most_common():
        ws3.append([name, cnt])

    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()


def write_llm_workbook_path(path: Path, headers: list[str], merged_rows: list[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_bytes(write_llm_workbook_bytes(headers, merged_rows))


def build_job_search_llm_xlsx_bytes(
    job: PipelineJob,
    *,
    batch_size: int = 40,
    max_workers: int | None = None,
) -> tuple[bytes, str]:
    qs = JdJobSearchRow.objects.filter(job=job).order_by("row_index")
    if not qs.exists():
        raise ValueError("该任务没有搜索结果数据，无法生成整理表")
    sk = nonempty_search_keys_for_job(job)
    sheaders = search_export_headers(job)
    rows_zh = [_search_row_csv_dict(obj, sk, sheaders) for obj in qs]
    by_index = run_llm_classify_by_row_index(
        rows_zh, batch_size=batch_size, max_workers=max_workers
    )
    header, merged = merge_llm_into_rows(sheaders, rows_zh, by_index)
    data = write_llm_workbook_bytes(header, merged)
    return data, f"job_{job.id}_search_llm_enriched.xlsx"


def build_from_csv_path_to_path(
    csv_path: Path,
    out_path: Path,
    *,
    batch_size: int = 40,
    max_workers: int | None = None,
    limit: int = 0,
) -> int:
    """CLI 用：读 pc_search_export.csv，写 xlsx；返回数据行数。"""
    import csv as csv_mod
    import io

    raw = csv_path.read_bytes()
    text: str
    for enc in ("utf-8-sig", "utf-8", "gb18030"):
        try:
            text = raw.decode(enc)
            break
        except UnicodeDecodeError:
            continue
    else:
        text = raw.decode("utf-8", errors="replace")
    if not text.strip():
        raise ValueError("CSV 无内容")
    reader = csv_mod.DictReader(io.StringIO(text))
    fieldnames = list(reader.fieldnames or [])
    rows = [dict(r) for r in reader]
    if limit and limit > 0:
        rows = rows[:limit]
    if not rows:
        raise ValueError("CSV 无有效数据行")
    by_index = run_llm_classify_by_row_index(
        rows, batch_size=batch_size, max_workers=max_workers
    )
    header, merged = merge_llm_into_rows(fieldnames, rows, by_index)
    write_llm_workbook_path(out_path, header, merged)
    return len(merged)

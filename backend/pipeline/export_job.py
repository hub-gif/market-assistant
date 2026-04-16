"""任务维度数据集导出：JSON / CSV（UTF-8 BOM）/ xlsx。列与源 CSV 对齐。"""
from __future__ import annotations

import csv
import json
from io import BytesIO, StringIO
from typing import Any

from django.db.models import QuerySet
from openpyxl import Workbook

from .csv_schema import (
    COMMENT_CSV_TO_FIELD,
    DETAIL_CSV_TO_FIELD,
    JD_SEARCH_CSV_HEADERS,
    MERGED_FIELD_TO_CSV_HEADER,
)
from .dataset_nonempty import (
    MATRIX_GROUP_COLUMN,
    comment_export_headers,
    detail_export_headers,
    merged_export_headers,
    nonempty_comment_fields_for_job,
    nonempty_detail_fields_for_job,
    nonempty_merged_fields_for_job,
    nonempty_search_keys_for_job,
    search_export_headers,
)
from .models import JdJobCommentRow, JdJobDetailRow, JdJobMergedRow, JdJobSearchRow, PipelineJob
from .row_serialize import (
    comment_row_to_dict,
    detail_row_to_dict,
    merged_row_to_dict,
    search_row_to_dict,
)


def _search_row_csv_dict(
    r: JdJobSearchRow, internal_keys: list[str], headers: list[str]
) -> dict[str, Any]:
    d = search_row_to_dict(r)
    out: dict[str, Any] = {"id": d["id"], "row_index": d["row_index"]}
    for k in internal_keys:
        out[JD_SEARCH_CSV_HEADERS[k]] = d.get(k, "")
    zh = MATRIX_GROUP_COLUMN["label"]
    if zh in headers:
        out[zh] = d.get("matrix_group_label", "")
    return out


def _detail_row_csv_dict(
    r: JdJobDetailRow, csv_cols: list[str], headers: list[str]
) -> dict[str, Any]:
    d = detail_row_to_dict(r)
    out: dict[str, Any] = {"id": d["id"], "row_index": d["row_index"]}
    for col in csv_cols:
        fn = DETAIL_CSV_TO_FIELD[col]
        out[col] = d.get(fn, "")
    zh = MATRIX_GROUP_COLUMN["label"]
    if zh in headers:
        out[zh] = d.get("matrix_group_label", "")
    return out


def _comment_row_csv_dict(r: JdJobCommentRow, csv_cols: list[str]) -> dict[str, Any]:
    d = comment_row_to_dict(r)
    out: dict[str, Any] = {"id": d["id"], "row_index": d["row_index"]}
    for col in csv_cols:
        fn = COMMENT_CSV_TO_FIELD[col]
        out[col] = d.get(fn, "")
    return out


def _merged_row_csv_dict(
    r: JdJobMergedRow, internal_keys: list[str], headers: list[str]
) -> dict[str, Any]:
    d = merged_row_to_dict(r)
    out: dict[str, Any] = {"id": d["id"], "row_index": d["row_index"]}
    for k in internal_keys:
        out[MERGED_FIELD_TO_CSV_HEADER[k]] = d.get(k, "")
    zh = MATRIX_GROUP_COLUMN["label"]
    if zh in headers:
        out[zh] = d.get("matrix_group_label", "")
    return out


def _prune_search_dict(d: dict[str, Any], internal_keys: list[str]) -> dict[str, Any]:
    out = {"id": d["id"], "row_index": d["row_index"]}
    for k in internal_keys:
        out[k] = d.get(k, "")
    return out


def _prune_detail_dict(d: dict[str, Any], fields: list[str]) -> dict[str, Any]:
    out = {"id": d["id"], "row_index": d["row_index"]}
    for k in fields:
        out[k] = d.get(k, "")
    return out


def _prune_comment_dict(d: dict[str, Any], fields: list[str]) -> dict[str, Any]:
    out = {"id": d["id"], "row_index": d["row_index"]}
    for k in fields:
        out[k] = d.get(k, "")
    return out


def _prune_merged_dict(d: dict[str, Any], fields: list[str]) -> dict[str, Any]:
    out = {"id": d["id"], "row_index": d["row_index"]}
    for k in fields:
        out[k] = d.get(k, "")
    return out


def _rows_as_list_search(job: PipelineJob) -> list[dict[str, Any]]:
    keys = nonempty_search_keys_for_job(job)
    extra_mg = JdJobSearchRow.objects.filter(job=job).exclude(matrix_group_label="").exists()
    qs = JdJobSearchRow.objects.filter(job=job)
    out: list[dict[str, Any]] = []
    for obj in qs.order_by("row_index").iterator(chunk_size=400):
        d = search_row_to_dict(obj)
        row = _prune_search_dict(d, keys)
        if extra_mg:
            row["matrix_group_label"] = d.get("matrix_group_label", "")
        out.append(row)
    return out


def _rows_as_list_detail(job: PipelineJob) -> list[dict[str, Any]]:
    fields = nonempty_detail_fields_for_job(job)
    extra_mg = JdJobDetailRow.objects.filter(job=job).exclude(matrix_group_label="").exists()
    qs = JdJobDetailRow.objects.filter(job=job)
    out: list[dict[str, Any]] = []
    for obj in qs.order_by("row_index").iterator(chunk_size=400):
        d = detail_row_to_dict(obj)
        row = _prune_detail_dict(d, fields)
        if extra_mg:
            row["matrix_group_label"] = d.get("matrix_group_label", "")
        out.append(row)
    return out


def _rows_as_list_comment(job: PipelineJob) -> list[dict[str, Any]]:
    fields = nonempty_comment_fields_for_job(job)
    qs = JdJobCommentRow.objects.filter(job=job)
    return [
        _prune_comment_dict(comment_row_to_dict(obj), fields)
        for obj in qs.order_by("row_index").iterator(chunk_size=400)
    ]


def _rows_as_list_merged(job: PipelineJob) -> list[dict[str, Any]]:
    fields = nonempty_merged_fields_for_job(job)
    extra_mg = JdJobMergedRow.objects.filter(job=job).exclude(matrix_group_label="").exists()
    qs = JdJobMergedRow.objects.filter(job=job)
    out: list[dict[str, Any]] = []
    for obj in qs.order_by("row_index").iterator(chunk_size=400):
        d = merged_row_to_dict(obj)
        row = _prune_merged_dict(d, fields)
        if extra_mg:
            row["matrix_group_label"] = d.get("matrix_group_label", "")
        out.append(row)
    return out


def build_json_bytes(*, job: PipelineJob, kind: str) -> tuple[bytes, str]:
    if kind == "search":
        data = _rows_as_list_search(job)
        name = f"job_{job.id}_search.json"
    elif kind == "detail":
        data = _rows_as_list_detail(job)
        name = f"job_{job.id}_detail.json"
    elif kind == "comments":
        data = _rows_as_list_comment(job)
        name = f"job_{job.id}_comments.json"
    elif kind == "all":
        data = {
            "job_id": job.id,
            "keyword": job.keyword,
            "search": _rows_as_list_search(job),
            "detail": _rows_as_list_detail(job),
            "comments": _rows_as_list_comment(job),
            "merged": _rows_as_list_merged(job),
        }
        name = f"job_{job.id}_all.json"
    elif kind == "merged":
        data = _rows_as_list_merged(job)
        name = f"job_{job.id}_merged.json"
    else:
        raise ValueError(f"unknown kind: {kind}")
    raw = json.dumps(data, ensure_ascii=False, indent=2)
    return raw.encode("utf-8"), name


def _write_csv_from_qs(
    *,
    qs: QuerySet,
    headers: list[str],
    row_fn: Any,
) -> str:
    buf = StringIO()
    w = csv.DictWriter(buf, fieldnames=headers, extrasaction="ignore")
    w.writeheader()
    for obj in qs.order_by("row_index").iterator(chunk_size=400):
        w.writerow(row_fn(obj))
    return buf.getvalue()


def build_csv_bytes(*, job: PipelineJob, kind: str) -> tuple[bytes, str]:
    if kind == "search":
        sk = nonempty_search_keys_for_job(job)
        headers = search_export_headers(job)
        text = _write_csv_from_qs(
            qs=JdJobSearchRow.objects.filter(job=job),
            headers=headers,
            row_fn=lambda o, _sk=sk, _h=headers: _search_row_csv_dict(o, _sk, _h),
        )
        name = f"job_{job.id}_search.csv"
    elif kind == "detail":
        headers = detail_export_headers(job)
        zh = MATRIX_GROUP_COLUMN["label"]
        dcols = [c for c in headers if c not in ("id", "row_index", zh)]
        text = _write_csv_from_qs(
            qs=JdJobDetailRow.objects.filter(job=job),
            headers=headers,
            row_fn=lambda o, _dc=dcols, _h=headers: _detail_row_csv_dict(o, _dc, _h),
        )
        name = f"job_{job.id}_detail.csv"
    elif kind == "comments":
        ccols = [c for c in comment_export_headers(job) if c not in ("id", "row_index")]
        text = _write_csv_from_qs(
            qs=JdJobCommentRow.objects.filter(job=job),
            headers=comment_export_headers(job),
            row_fn=lambda o, _cc=ccols: _comment_row_csv_dict(o, _cc),
        )
        name = f"job_{job.id}_comments.csv"
    elif kind == "all":
        sk = nonempty_search_keys_for_job(job)
        sheaders = search_export_headers(job)
        dheaders = detail_export_headers(job)
        zh = MATRIX_GROUP_COLUMN["label"]
        dcols = [c for c in dheaders if c not in ("id", "row_index", zh)]
        ccols = [c for c in comment_export_headers(job) if c not in ("id", "row_index")]
        mk = nonempty_merged_fields_for_job(job)
        mheaders = merged_export_headers(job)
        parts = [
            "# search",
            _write_csv_from_qs(
                qs=JdJobSearchRow.objects.filter(job=job),
                headers=sheaders,
                row_fn=lambda o, _sk=sk, _h=sheaders: _search_row_csv_dict(o, _sk, _h),
            ),
            "",
            "# detail",
            _write_csv_from_qs(
                qs=JdJobDetailRow.objects.filter(job=job),
                headers=dheaders,
                row_fn=lambda o, _dc=dcols, _h=dheaders: _detail_row_csv_dict(
                    o, _dc, _h
                ),
            ),
            "",
            "# comments",
            _write_csv_from_qs(
                qs=JdJobCommentRow.objects.filter(job=job),
                headers=comment_export_headers(job),
                row_fn=lambda o, _cc=ccols: _comment_row_csv_dict(o, _cc),
            ),
            "",
            "# merged",
            _write_csv_from_qs(
                qs=JdJobMergedRow.objects.filter(job=job),
                headers=mheaders,
                row_fn=lambda o, _mk=mk, _h=mheaders: _merged_row_csv_dict(o, _mk, _h),
            ),
        ]
        text = "\n".join(parts)
        name = f"job_{job.id}_all.csv"
    elif kind == "merged":
        mk = nonempty_merged_fields_for_job(job)
        headers = merged_export_headers(job)
        text = _write_csv_from_qs(
            qs=JdJobMergedRow.objects.filter(job=job),
            headers=headers,
            row_fn=lambda o, _mk=mk, _h=headers: _merged_row_csv_dict(o, _mk, _h),
        )
        name = f"job_{job.id}_merged.csv"
    else:
        raise ValueError(f"unknown kind: {kind}")
    return ("\ufeff" + text).encode("utf-8"), name


def _append_sheet(ws, headers: list[str], qs: QuerySet, row_fn: Any) -> None:
    ws.append(headers)
    for obj in qs.order_by("row_index").iterator(chunk_size=400):
        rowd = row_fn(obj)
        ws.append([rowd.get(h, "") for h in headers])


def build_xlsx_bytes(*, job: PipelineJob, kind: str) -> tuple[bytes, str]:
    wb = Workbook()
    if kind == "search":
        ws = wb.active
        ws.title = "search"[:31]
        sk = nonempty_search_keys_for_job(job)
        sheaders = search_export_headers(job)
        _append_sheet(
            ws,
            sheaders,
            JdJobSearchRow.objects.filter(job=job),
            lambda o, _sk=sk, _h=sheaders: _search_row_csv_dict(o, _sk, _h),
        )
        name = f"job_{job.id}_search.xlsx"
    elif kind == "detail":
        ws = wb.active
        ws.title = "detail"[:31]
        dheaders = detail_export_headers(job)
        zh = MATRIX_GROUP_COLUMN["label"]
        dcols = [c for c in dheaders if c not in ("id", "row_index", zh)]
        _append_sheet(
            ws,
            dheaders,
            JdJobDetailRow.objects.filter(job=job),
            lambda o, _dc=dcols, _h=dheaders: _detail_row_csv_dict(o, _dc, _h),
        )
        name = f"job_{job.id}_detail.xlsx"
    elif kind == "comments":
        ws = wb.active
        ws.title = "comments"[:31]
        ccols = [c for c in comment_export_headers(job) if c not in ("id", "row_index")]
        _append_sheet(
            ws,
            comment_export_headers(job),
            JdJobCommentRow.objects.filter(job=job),
            lambda o, _cc=ccols: _comment_row_csv_dict(o, _cc),
        )
        name = f"job_{job.id}_comments.xlsx"
    elif kind == "all":
        sk = nonempty_search_keys_for_job(job)
        sheaders = search_export_headers(job)
        dheaders = detail_export_headers(job)
        zh = MATRIX_GROUP_COLUMN["label"]
        dcols = [c for c in dheaders if c not in ("id", "row_index", zh)]
        ccols = [c for c in comment_export_headers(job) if c not in ("id", "row_index")]
        mk = nonempty_merged_fields_for_job(job)
        mheaders = merged_export_headers(job)
        ws1 = wb.active
        ws1.title = "search"[:31]
        _append_sheet(
            ws1,
            sheaders,
            JdJobSearchRow.objects.filter(job=job),
            lambda o, _sk=sk, _h=sheaders: _search_row_csv_dict(o, _sk, _h),
        )
        ws2 = wb.create_sheet("detail"[:31])
        _append_sheet(
            ws2,
            dheaders,
            JdJobDetailRow.objects.filter(job=job),
            lambda o, _dc=dcols, _h=dheaders: _detail_row_csv_dict(o, _dc, _h),
        )
        ws3 = wb.create_sheet("comments"[:31])
        _append_sheet(
            ws3,
            comment_export_headers(job),
            JdJobCommentRow.objects.filter(job=job),
            lambda o, _cc=ccols: _comment_row_csv_dict(o, _cc),
        )
        ws4 = wb.create_sheet("merged"[:31])
        _append_sheet(
            ws4,
            mheaders,
            JdJobMergedRow.objects.filter(job=job),
            lambda o, _mk=mk, _h=mheaders: _merged_row_csv_dict(o, _mk, _h),
        )
        name = f"job_{job.id}_all.xlsx"
    elif kind == "merged":
        mk = nonempty_merged_fields_for_job(job)
        ws = wb.active
        ws.title = "merged"[:31]
        mheaders = merged_export_headers(job)
        _append_sheet(
            ws,
            mheaders,
            JdJobMergedRow.objects.filter(job=job),
            lambda o, _mk=mk, _h=mheaders: _merged_row_csv_dict(o, _mk, _h),
        )
        name = f"job_{job.id}_merged.xlsx"
    else:
        raise ValueError(f"unknown kind: {kind}")
    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue(), name
